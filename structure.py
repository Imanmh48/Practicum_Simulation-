import random
import os

from sim1 import VolunteerMetrics
from Event import Event
from config import *
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import PatternFill
import pandas as pd
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties

#random.seed(1)
class Participant:
    def __init__(self, name, base_score, personality):
        self.name = name
        self.base_score = base_score

        self.response_time = 0
        self.attendance_rate = 0
        self.task_completion_rate = 0

        self.event_score = 0
        self.metrics_score = 0 # the current postion of the participant in the metrics

        self.hours_commitment = 0
        self.team_performance = 0
        self.problem_solving = 0
        self.conflict_resolution = 0
        self.leadership_metrics = 0
        self.leadership_appointments = 0

        self.total_score = base_score
        self.rank = ""
        self.inactivity_period = 0
        self.personality = personality

    def calculate_event_score(self, event_size):
        # If participant is inactive, event_score is 0
        if self.inactivity_period >= 1:
            self.event_score = 0
            return
        
        # Ensure minimum event size
        if event_size < 10:
            raise ValueError(f"Event size cannot be less than 10 (got {event_size})")

        # Direct threshold checks from largest to smallest
        if event_size >= 200:
            self.event_score = 50
        elif event_size >= 100:
            self.event_score = 100
        elif event_size >= 50:
            self.event_score = 150
        else:  # event_size <= 50
            self.event_score = 200
    
    def update_metrics_score(self, event, response_time_mins, late_arrivals, early_departures, 
                           unscheduled_absences, completed_tasks, total_tasks,
                           logged_hours, expected_hours, team_completed, team_total,
                           actual_time, planned_time, problem_time, expected_problem_time,
                           successful_solutions, total_solutions, conflicts_resolved, 
                           total_conflicts):
        """Update metrics score based on event's metrics weights and provided metric values"""
        # Get weights from event object
        weights = event._get_weights_for_type(event.event_type)
        
        # Calculate new metric scores with weights applied
        metrics = VolunteerMetrics()
        
        # Calculate new metric scores with weights applied
        new_response_time = metrics.calculate_response_time(response_time_mins) * weights['response_time']
        new_attendance = metrics.calculate_attendance(late_arrivals, early_departures, unscheduled_absences, 5) * weights['attendance_rate']
        new_task_completion = metrics.calculate_task_completion(completed_tasks, total_tasks) * weights['task_completion']
        new_hours = metrics.calculate_hours_commitment(logged_hours, expected_hours)  # No weight for hours
        new_team_perf = metrics.calculate_team_performance(team_completed, team_total, actual_time, planned_time) * weights['team_performance']
        new_problem_solving = metrics.calculate_problem_solving(problem_time, expected_problem_time, successful_solutions, total_solutions) * weights['problem_solving']
        new_conflict = metrics.calculate_conflict_resolution(conflicts_resolved, total_conflicts) * weights['conflict_resolution']
        new_leadership = metrics.calculate_leadership_metrics(team_completed, team_total, self.leadership_appointments) * weights['leadership']
        
        # Update metrics by averaging with previous weighted values
        self.response_time = (self.response_time + new_response_time) / 2
        self.attendance_rate = (self.attendance_rate + new_attendance) / 2
        self.task_completion_rate = (self.task_completion_rate + new_task_completion) / 2
        self.hours_commitment = (self.hours_commitment + new_hours) / 2
        self.team_performance = (self.team_performance + new_team_perf) / 2
        self.problem_solving = (self.problem_solving + new_problem_solving) / 2
        self.conflict_resolution = (self.conflict_resolution + new_conflict) / 2
        self.leadership_metrics = (self.leadership_metrics + new_leadership) / 2
        
        # Sum up all weighted metrics for final score
        self.metrics_score = (
            self.response_time +
            self.attendance_rate + 
            self.task_completion_rate +
            self.team_performance +
            self.problem_solving +
            self.leadership_metrics +
            self.conflict_resolution
        )

    def determine_rank(self, thresholds):
        # Make Platinum more exclusive - only top performers get it
        if self.total_score >= thresholds[0] and self.metrics_score >= PLATINUM_METRICS_THRESHOLD:  # Requiring high metrics performance
            self.rank = 'Platinum'
        elif self.total_score >= thresholds[1]:
            self.rank = 'Gold'
        elif self.total_score >= thresholds[2]:
            self.rank = 'Silver'
        elif self.total_score >= thresholds[3]:
            self.rank = 'Bronze'
        else:
            self.rank = 'Bronze'
    

    def apply_decay(self, inactivity_threshold=2):
        # Calculate the current total score first
        previous_metrics = getattr(self, '_previous_metrics_score', self.metrics_score)
        if self.inactivity_period >= 1 or (hasattr(self, '_previous_metrics_score') and (previous_metrics - self.metrics_score) > METRICS_DECLINE_THRESHOLD):
            current_total = self.total_score
        else:
            metrics_modifier = 1 + (self.metrics_score / 50)
            base_value = 100 if self.base_score == 0 else self.base_score
            current_total = (base_value + self.event_score) * metrics_modifier

        total_score_before_decay = current_total

        # Apply inactivity decay if inactivity_period equals threshold
        if self.inactivity_period == inactivity_threshold:
            if self.rank in DECAY_RATES:
                print(f"Total Score before decay for {self.name}: {total_score_before_decay:.2f}")
                inactivity_decay = total_score_before_decay * DECAY_RATES[self.rank]
                total_score_before_decay -= inactivity_decay
                print(f"Decay applied for inactivity to {self.name} ({self.rank}): -{inactivity_decay:.2f}")
                # Reset inactivity period after applying decay
                self.inactivity_period = 0

        # Apply metrics decline decay using total_score_before_decay
        previous_metrics = getattr(self, '_previous_metrics_score', self.metrics_score)
        original_score = total_score_before_decay
        if hasattr(self, '_previous_metrics_score') and (previous_metrics - self.metrics_score) > METRICS_DECLINE_THRESHOLD:
            metrics_decay = original_score * METRICS_DECLINE_DECAY_RATE
            total_score_before_decay -= metrics_decay
            print(f"Total Score before metrics decay: {original_score:.2f}")
            print(f"Decay applied for metrics decline to {self.name}: -{metrics_decay:.2f} (Metrics dropped from {previous_metrics:.2f} to {self.metrics_score:.2f})")
        
        # Store current metrics score for next comparison
        self._previous_metrics_score = self.metrics_score
        
        # Set the final score
        self.total_score = total_score_before_decay

def apply_reset(list_participants,rank_distribution): #this will apply the reset for all classes

    plat=rank_distribution[0]
    gold=rank_distribution[1]
    silver=rank_distribution[2]
    for participant in list_participants:
        if participant.base_score>=plat: #will change this to be if they are above platinum they will be reset to platinum
            participant.base_score=plat
        elif participant.base_score>=gold:
            participant.base_score-=100 #demote them by 2 tiers
            #print("In gold:\t"+participant.name+"\t"+str(participant.base_score))
        elif participant.base_score>=silver:
            participant.base_score-=50 #demote by 1 tier
            #print("In silver:\t"+participant.name+"\t"+str(participant.base_score))
        else:
            participant.base_score-=20 #demote by a small amount because they are already doing bad
            #print("In bronze:\t"+participant.name+"\t"+str(participant.base_score))


def prepare_distributed_reset(event_sizes):
    return NUMBER_OF_SEASONS, distrubte_events_across_seasons(event_sizes, NUMBER_OF_SEASONS), 0

def distrubte_events_across_seasons(num_of_events,num_of_seasons):
    counter=0
    while True: 
        event_assignment=[]
        for i in range(num_of_seasons):
            num_of_event=random.randint(1,num_of_events) #generate a random number of events
            event_assignment.append(num_of_event) # append that number
        if sum(event_assignment)==num_of_events: # make sure that my generated number of event is equal to all_events
            break
        counter+=1
        if counter==2000: #safeguard
            print("out")
            break
    return event_assignment

def create_excel_graphs(participants_history, threshold_type, event_numbers, wb=None):
    """
    Create Excel graphs from simulation data with improved visual styling
    """
    if wb is None:
        wb = Workbook()
    
    # Overall Progress Sheet
    ws_overall = wb.create_sheet(f"Overall Progress ({threshold_type})")
    
    # Write data in a hidden sheet
    data_sheet = wb.create_sheet(f"Data_{threshold_type}")
    data_sheet.sheet_state = 'hidden'
    
    # Write headers to hidden sheet
    data_sheet.cell(row=1, column=1, value="Event")
    for col, participant in enumerate(participants_history.keys(), start=2):
        data_sheet.cell(row=1, column=col, value=participant)
    
    # Write data to hidden sheet
    for row, event_num in enumerate(event_numbers, start=2):
        data_sheet.cell(row=row, column=1, value=event_num)
        for col, scores in enumerate(participants_history.values(), start=2):
            data_sheet.cell(row=row, column=col, value=scores[row-2])
    
    # Create overall line chart
    chart = LineChart()
    chart.title = f"Volunteer Progress ({threshold_type.capitalize()} Thresholds)"
    chart.style = 13
    chart.x_axis.title = "Event Number"
    chart.y_axis.title = "Total Score"
    chart.height = 15
    chart.width = 20
    
    data = Reference(data_sheet, min_col=2, min_row=1, 
                    max_col=len(participants_history)+1, 
                    max_row=len(event_numbers)+1)
    cats = Reference(data_sheet, min_col=1, min_row=2, 
                    max_row=len(event_numbers)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Style lines with distinct colors
    colors = ['FF0000', '00FF00', '0000FF', 'FFA500', '800080', 
              '008080', 'FF69B4', '4B0082', '006400', 'DC143C', 
              '00FFFF', '8B4513', 'FF8C00', '9400D3', '808000', 
              'FF1493', 'CD853F', '00FA9A', 'FFD700', '4682B4']
    
    for i, series in enumerate(chart.series):
        series.smooth = True
        series.graphicalProperties.line.width = 25000
        series.graphicalProperties.line.solidFill = colors[i % len(colors)]
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = colors[i % len(colors)]
        series.marker.graphicalProperties.line.solidFill = colors[i % len(colors)]
    
    # Position legend for overall progress chart
    chart.legend.position = 'r'
    chart.plot_area.layout = Layout(
        manualLayout=ManualLayout(
            x=0.1,    # Left position
            y=0.1,    # Top position
            h=0.8,    # Height
            w=0.7     # Width - reduced to make room for legend
        )
    )
    
    ws_overall.add_chart(chart, "A1")
    
    # Personality-based Sheet
    ws_personality = wb.create_sheet(f"Personality Based ({threshold_type})")
    personality_data = {}
    
    # Group participants by personality
    for participant in participants:
        if participant.personality not in personality_data:
            personality_data[participant.personality] = []
        personality_data[participant.personality].append(participants_history[participant.name])
    
    # Calculate averages for each personality type
    personality_sheet = wb.create_sheet(f"Personality_Data_{threshold_type}")
    personality_sheet.sheet_state = 'hidden'
    
    row = 1
    personality_sheet.cell(row=row, column=1, value="Event")
    for col, personality in enumerate(personality_data.keys(), start=2):
        personality_sheet.cell(row=row, column=col, value=personality)
    
    for event_idx, event_num in enumerate(event_numbers):
        row = event_idx + 2
        personality_sheet.cell(row=row, column=1, value=event_num)
        for col, (personality, scores_list) in enumerate(personality_data.items(), start=2):
            avg_score = sum(score_list[event_idx] for score_list in scores_list) / len(scores_list)
            personality_sheet.cell(row=row, column=col, value=avg_score)
    
    # Create personality chart
    personality_chart = LineChart()
    personality_chart.title = f"Average Scores by Personality ({threshold_type.capitalize()} Thresholds)"
    personality_chart.style = 13
    personality_chart.x_axis.title = "Event Number"
    personality_chart.y_axis.title = "Average Score"
    personality_chart.height = 15
    personality_chart.width = 20
    
    data = Reference(personality_sheet, min_col=2, min_row=1, 
                    max_col=len(personality_data)+1, 
                    max_row=len(event_numbers)+1)
    cats = Reference(personality_sheet, min_col=1, min_row=2, 
                    max_row=len(event_numbers)+1)
    
    personality_chart.add_data(data, titles_from_data=True)
    personality_chart.set_categories(cats)
    
    # Style personality chart lines
    personality_colors = {
        'lazy': 'FF0000',      # Red
        'ideal': '00FF00',     # Green
        'inconsistent': 'FFA500', # Orange
        'growing': '0000FF',   # Blue
        'average': '800080'    # Purple
    }
    
    for i, series in enumerate(personality_chart.series):
        personality = list(personality_data.keys())[i]
        series.smooth = True
        series.graphicalProperties.line.width = 25000
        series.graphicalProperties.line.solidFill = personality_colors[personality]
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = personality_colors[personality]
        series.marker.graphicalProperties.line.solidFill = personality_colors[personality]
    
    # Position legend for personality chart
    personality_chart.legend.position = 'r'
    personality_chart.plot_area.layout = Layout(
        manualLayout=ManualLayout(
            x=0.1,    # Left position
            y=0.1,    # Top position
            h=0.8,    # Height
            w=0.7     # Width - reduced to make room for legend
        )
    )
    
    ws_personality.add_chart(personality_chart, "A1")
    
    return wb

def create_rank_progression_graphs(participants_history, rank_history, threshold_type, event_numbers, wb=None):
    if wb is None:
        wb = Workbook()
    
    # Define rank mapping
    rank_mapping = {
        'Platinum': 1,
        'Gold': 2,
        'Silver': 3,
        'Bronze': 4,
        'None': 5  # Assuming 'None' is the lowest rank
    }
    
    # Overall Rank Progression Sheet
    ws_rank_progression = wb.create_sheet(f"Rank Progression ({threshold_type})")
    
    # Write data in a hidden sheet
    data_sheet = wb.create_sheet(f"Rank_Data_{threshold_type}")
    data_sheet.sheet_state = 'hidden'
    
    # Write headers to hidden sheet
    data_sheet.cell(row=1, column=1, value="Event")
    for col, participant in enumerate(participants_history.keys(), start=2):
        data_sheet.cell(row=1, column=col, value=participant)
    
    # Write data to hidden sheet
    for row, event_num in enumerate(event_numbers, start=2):
        data_sheet.cell(row=row, column=1, value=event_num)
        for col, ranks in enumerate(rank_history.values(), start=2):
            data_sheet.cell(row=row, column=col, value=rank_mapping[ranks[row-2]])
    
    # Create overall rank progression line chart
    chart = LineChart()
    chart.title = f"Volunteer Rank Progression ({threshold_type.capitalize()} Thresholds)"
    chart.style = 13
    chart.x_axis.title = "Event Number"
    chart.y_axis.title = "Rank"
    chart.height = 15
    chart.width = 20
    
    data = Reference(data_sheet, min_col=2, min_row=1, 
                    max_col=len(participants_history)+1, 
                    max_row=len(event_numbers)+1)
    cats = Reference(data_sheet, min_col=1, min_row=2, 
                    max_row=len(event_numbers)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Style lines with distinct colors
    colors = ['FF0000', '00FF00', '0000FF', 'FFA500', '800080', 
              '008080', 'FF69B4', '4B0082', '006400', 'DC143C']
    
    for i, series in enumerate(chart.series):
        series.smooth = True
        series.graphicalProperties.line.width = 25000
        series.graphicalProperties.line.solidFill = colors[i % len(colors)]
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = colors[i % len(colors)]
        series.marker.graphicalProperties.line.solidFill = colors[i % len(colors)]
    
    # Position legend
    chart.legend.position = 'r'
    chart.legend.layout = Layout(
        manualLayout=ManualLayout(
            x=1.0,
            y=0.5,
            h=0.9,
            w=0.2
        )
    )
    
    ws_rank_progression.add_chart(chart, "A1")
    
    # Personality-based Rank Progression Sheet
    ws_personality_rank = wb.create_sheet(f"Personality Rank ({threshold_type})")
    personality_data = {}
    
    # Group participants by personality
    for participant in participants:
        if participant.personality not in personality_data:
            personality_data[participant.personality] = []
        personality_data[participant.personality].append(rank_history[participant.name])
    
    # Calculate averages for each personality type
    personality_sheet = wb.create_sheet(f"Personality_Rank_Data_{threshold_type}")
    personality_sheet.sheet_state = 'hidden'
    
    row = 1
    personality_sheet.cell(row=row, column=1, value="Event")
    for col, personality in enumerate(personality_data.keys(), start=2):
        personality_sheet.cell(row=row, column=col, value=personality)
    
    for event_idx, event_num in enumerate(event_numbers):
        row = event_idx + 2
        personality_sheet.cell(row=row, column=1, value=event_num)
        for col, (personality, ranks_list) in enumerate(personality_data.items(), start=2):
            avg_rank = sum(rank_mapping[rank_list[event_idx]] for rank_list in ranks_list) / len(ranks_list)
            personality_sheet.cell(row=row, column=col, value=avg_rank)
    
    # Create personality rank chart
    personality_chart = LineChart()
    personality_chart.title = f"Average Rank by Personality ({threshold_type.capitalize()} Thresholds)"
    personality_chart.style = 13
    personality_chart.x_axis.title = "Event Number"
    personality_chart.y_axis.title = "Average Rank"
    personality_chart.height = 15
    personality_chart.width = 20
    
    data = Reference(personality_sheet, min_col=2, min_row=1, 
                    max_col=len(personality_data)+1, 
                    max_row=len(event_numbers)+1)
    cats = Reference(personality_sheet, min_col=1, min_row=2, 
                    max_row=len(event_numbers)+1)
    
    personality_chart.add_data(data, titles_from_data=True)
    personality_chart.set_categories(cats)
    
    # Style personality chart lines
    personality_colors = {
        'lazy': 'FF0000',      # Red
        'ideal': '00FF00',     # Green
        'inconsistent': 'FFA500', # Orange
        'growing': '0000FF',   # Blue
        'average': '800080'    # Purple
    }
    
    for i, series in enumerate(personality_chart.series):
        personality = list(personality_data.keys())[i]
        series.smooth = True
        series.graphicalProperties.line.width = 25000
        series.graphicalProperties.line.solidFill = personality_colors[personality]
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = personality_colors[personality]
        series.marker.graphicalProperties.line.solidFill = personality_colors[personality]
    
    personality_chart.legend.position = 'r'
    personality_chart.legend.layout = Layout(
        manualLayout=ManualLayout(
            x=1.0,
            y=0.5,
            h=0.9,
            w=0.2
        )
    )
    
    ws_personality_rank.add_chart(personality_chart, "A1")
    
    return wb

def simulate_events(num_events, event_size, participants, start_event_number, thresholds, inactivity_threshold=3, threshold_type="standard"):
    print(f"\nSimulating with Event Size: {event_size}")
    print("=" * 80)
    
    # Initialize history tracking
    participants_history = {p.name: [] for p in participants}
    rank_history = {p.name: [] for p in participants}
    event_numbers = []
    
    for event_number in range(start_event_number, start_event_number + num_events):
        event_numbers.append(event_number)
        
        print(f"\nEvent {event_number} (Event Size: {event_size}):")
        print("=" * 50)
        print(f"{'Name':<10} | {'Base Score':<10} | {'Metrics_Score':<15} | {'Event Score':<15} | {'personality':<12}|{'Total Score':<15} | {'Rank':<6} | {'Inactivity':<10}")
        print("-" * 92)

        for participant in participants:
            # Store the previous total score before calculating new scores
            previous_total = participant.total_score
            
            # Track inactivity based on participant name and event number
            was_inactive = False
            participant_number = int(participant.name[1:])  # Extract number from "v1", "v2", etc.
            
            # Group 1 (v1-v5): Inactive during first 3 events AND every 5th event after that
            is_group_1 = (participant_number <= 5) and (event_number <= 3 or event_number % 5 == 0)
            
            # Group 2 (v6-v10): Inactive during events 4-7 AND every 4th event after that
            is_group_2 = (6 <= participant_number <= 10) and (4 <= event_number <= 7 or event_number % 4 == 0)
            
            # Group 3 (v11-v15): Inactive every third event AND for events 10-12
            is_group_3 = (11 <= participant_number <= 15) and (event_number % 3 == 0 or 10 <= event_number <= 12)
            
            # Group 4 (v16-v20): Inactive during events 8-10 AND every 6th event
            is_group_4 = (16 <= participant_number <= 20) and (8 <= event_number <= 10 or event_number % 6 == 0)
            
            # Special condition: All participants except v1, v7, v13, and v19 are inactive every 8th event
            is_special_inactive = participant_number not in [1, 7, 13, 19] and event_number % 8 == 0
            
            if is_group_1 or is_group_2 or is_group_3 or is_group_4 or is_special_inactive:
                participant.inactivity_period += 1
                was_inactive = True
            else:
                participant.inactivity_period = 0

            event = Event("Test Event", event_size, "2024-01-01", "standard")
            
            # Generate much more varied random values for each participant
            if participant.personality == "lazy":
                response_time = random.randint(80, 121)      # Very poor response time
                late_arrivals = random.randint(3, 6)        # Many late arrivals
                completed_tasks = random.randint(1, 4)       # Very poor completion
                team_completed = random.randint(1, 4)        # Very poor team performance
                successful_solutions = random.randint(1, 4)   # Very poor problem solving
                conflicts_resolved = random.randint(1, 4)    # Very poor conflict resolution
            elif participant.personality == "ideal":
                response_time = random.randint(5, 16)       # Excellent response time
                late_arrivals = 0                           # No late arrivals
                completed_tasks = random.randint(9, 11)      # Nearly perfect completion
                successful_solutions = random.randint(9, 11)  # Excellent problem solving
                conflicts_resolved = random.randint(9, 11)   # Excellent conflict resolution
            elif participant.personality == "inconsistent":
                response_time = random.randint(15,60)
                late_arrivals = random.randint(0,4)
                completed_tasks = random.randint(5,11)
                successful_solutions = random.randint(5,11)
                conflicts_resolved = random.randint(5,11)
            elif participant.personality == "growing":
                response_time = random.randint(0,round(participant.response_time))
                late_arrivals = 0
                completed_tasks = random.randint(round(participant.task_completion_rate),11)
                successful_solutions = random.randint(round(participant.task_completion_rate),11)
                conflicts_resolved = random.randint(round(participant.conflict_resolution),11)
            elif participant.personality == "average":
                response_time = random.randint(35,65)
                late_arrivals = random.randint(0,3)
                completed_tasks = random.randint(4,8)
                successful_solutions = random.randint(4,8)
                conflicts_resolved = random.randint(4,8)
        
            # More variable common random values
            early_departures = random.randint(0, 2)
            unscheduled_absences = random.randint(0, 2)
            total_tasks = 10
            logged_hours = random.randint(25, 40)          # More variable hours
            expected_hours = 40
            team_total = 10
            actual_time = random.randint(25, 40)           # More variable time management
            planned_time = 40
            problem_time = random.randint(25, 40)          # More variable problem solving
            expected_problem_time = 40
            total_solutions = 10
            total_conflicts = 10
            
            # Only update metrics if participant is active
            if not was_inactive:
                participant.update_metrics_score(
                    event=event,
                    response_time_mins=response_time,
                    late_arrivals=late_arrivals,
                    early_departures=early_departures,
                    unscheduled_absences=unscheduled_absences,
                    completed_tasks=completed_tasks,
                    total_tasks=total_tasks,
                    logged_hours=logged_hours,
                    expected_hours=expected_hours,
                    team_completed=team_completed,
                    team_total=team_total,
                    actual_time=actual_time,
                    planned_time=planned_time,
                    problem_time=problem_time,
                    expected_problem_time=expected_problem_time,
                    successful_solutions=successful_solutions,
                    total_solutions=total_solutions,
                    conflicts_resolved=conflicts_resolved,
                    total_conflicts=total_conflicts
                )
            
            participant.calculate_event_score(event_size)
            participant.determine_rank(thresholds)
            participant.apply_decay(inactivity_threshold=inactivity_threshold)
            participant.determine_rank(thresholds)
            
            # Update base_score to previous event's total score
            participant.base_score = previous_total

            # Track scores after all calculations
            participants_history[participant.name].append(participant.total_score)
            rank_history[participant.name].append(participant.rank)

            inactivity_display = f"{participant.inactivity_period} months" if participant.inactivity_period > 0 else "Active"
            
            print(f"{participant.name:<10} | {participant.base_score:<10.1f} | {participant.metrics_score:<15.2f} | {participant.event_score:<15.2f} |{participant.personality:<12} |{participant.total_score:<15.1f} | {participant.rank:<6} | {participant.inactivity_period:<10}")

        print("=" * 50)
    
    return participants_history, rank_history, event_numbers

switch_between_reset_modes=True # changing the reset methods used below
participants = []
personality_counter = 0
skip = NUMBER_OF_VOLUNTEERS//len(PERSONALITIES)
leftovers = NUMBER_OF_VOLUNTEERS%len(PERSONALITIES)
for batch in range(0,NUMBER_OF_VOLUNTEERS,skip): 
    if personality_counter<len(PERSONALITIES):
        for v in range(batch,batch+skip):
            participants.append(Participant("v"+str(v+1),INITIAL_BASE_SCORES["v"+str(v+1)],PERSONALITIES[personality_counter]))
        personality_counter+=1

if leftovers!=0:
    for leftover in range(leftovers):
        v+=1
        participants.append(Participant("v"+str(v+1),INITIAL_BASE_SCORES["v"+str(v+1)],PERSONALITIES[leftover]))



# Test different event sizes with continuous event numbering
shuffled_sizes = EVENT_SIZES.copy()

# At the start of the simulation loop, create one workbook
workbook = Workbook()
print("Created new workbook")

# Before the inactivity threshold loop
simulation_results = []

for inactivity_threshold in INACTIVITY_THRESHOLDS:
    print(f"\n\n{'='*50}")
    print(f"TESTING WITH INACTIVITY THRESHOLD = {inactivity_threshold}")
    print(f"{'='*50}\n")

    # Reset participants for new threshold test
    for participant in participants:
        participant.base_score = INITIAL_BASE_SCORES[participant.name]
        participant.total_score = participant.base_score
        participant.inactivity_period = 0

    # First simulation with standard thresholds
    current_event_number = 1
    all_participants_history = {p.name: [] for p in participants}
    all_rank_history = {p.name: [] for p in participants}
    all_event_numbers = []

    shuffled_sizes = EVENT_SIZES.copy()
    random.shuffle(shuffled_sizes)
    print(f"\nUsing standard thresholds: {THRESHOLDS['standard']}")
    print(f"Shuffled event sizes: {shuffled_sizes}")
    number_of_seasons,event_distribution,counter = prepare_distributed_reset(len(EVENT_SIZES))
    breakpoint = event_distribution[counter]
    if switch_between_reset_modes:
        print(event_distribution)
    print("Season", counter+1)
    print("#" * 90)     
    for event_size in shuffled_sizes:
        history, rank_history, event_nums = simulate_events(1, event_size, participants, current_event_number, 
                                           THRESHOLDS["standard"], inactivity_threshold, "standard")
        
        # Accumulate history
        for participant in participants:
            all_participants_history[participant.name].extend(history[participant.name])
            all_rank_history[participant.name].extend(rank_history[participant.name])
        all_event_numbers.extend(event_nums)
        
        if switch_between_reset_modes:
            if (current_event_number) == breakpoint:
                apply_reset(participants,THRESHOLDS["standard"])
                counter += 1
                try:
                    breakpoint += event_distribution[counter]
                    print("Season", counter+1)
                    print("#" * 90)
                except:
                    pass
        current_event_number += 1

    print(f"\nCreating graphs for standard thresholds (Inactivity: {inactivity_threshold})...")
    workbook = create_excel_graphs(all_participants_history, f"standard_{inactivity_threshold}", all_event_numbers, workbook)
    workbook = create_rank_progression_graphs(all_participants_history, all_rank_history, f"standard_{inactivity_threshold}", all_event_numbers, workbook)

    # Reset participants for competitive thresholds
    for participant in participants:
        participant.base_score = INITIAL_BASE_SCORES[participant.name]
        participant.total_score = participant.base_score
        participant.inactivity_period = 0

    current_event_number = 1
    all_participants_history = {p.name: [] for p in participants}
    all_rank_history = {p.name: [] for p in participants}
    all_event_numbers = []

    shuffled_sizes = EVENT_SIZES.copy()
    random.shuffle(shuffled_sizes)
    print(f"\nUsing competitive thresholds: {THRESHOLDS['competitive']}")
    print(f"Shuffled event sizes: {shuffled_sizes}")
    number_of_seasons,event_distribution,counter = prepare_distributed_reset(len(EVENT_SIZES))
    breakpoint = event_distribution[counter]
    if switch_between_reset_modes:
        print(event_distribution)
    counter=0
    for event_size in shuffled_sizes:
        history, rank_history, event_nums = simulate_events(1, event_size, participants, current_event_number, 
                                           THRESHOLDS["competitive"], inactivity_threshold, "competitive")
        
        # Accumulate history
        for participant in participants:
            all_participants_history[participant.name].extend(history[participant.name])
            all_rank_history[participant.name].extend(rank_history[participant.name])
        all_event_numbers.extend(event_nums)
        
        if switch_between_reset_modes:
            if (current_event_number) == breakpoint:
                apply_reset(participants,THRESHOLDS["competitive"])
                counter += 1
                try:
                    breakpoint += event_distribution[counter]
                    print("Season", counter+1)
                    print("#" * 90)
                except:
                    pass
        current_event_number += 1

    print(f"\nCreating graphs for competitive thresholds (Inactivity: {inactivity_threshold})...")
    workbook = create_excel_graphs(all_participants_history, f"competitive_{inactivity_threshold}", all_event_numbers, workbook)
    workbook = create_rank_progression_graphs(all_participants_history, all_rank_history, f"competitive_{inactivity_threshold}", all_event_numbers, workbook)

    # Reset participants for strict thresholds
    for participant in participants:
        participant.base_score = INITIAL_BASE_SCORES[participant.name]
        participant.total_score = participant.base_score
        participant.inactivity_period = 0

    current_event_number = 1
    all_participants_history = {p.name: [] for p in participants}
    all_rank_history = {p.name: [] for p in participants}
    all_event_numbers = []

    shuffled_sizes = EVENT_SIZES.copy()
    random.shuffle(shuffled_sizes)
    print(f"\nUsing strict thresholds: {THRESHOLDS['strict']}")
    print(f"Shuffled event sizes: {shuffled_sizes}")
    number_of_seasons,event_distribution,counter = prepare_distributed_reset(len(EVENT_SIZES))
    breakpoint = event_distribution[counter]
    if switch_between_reset_modes:
        print(event_distribution)
    print("Season", counter+1)
    print("#" * 90)
    for event_size in shuffled_sizes:
        history, rank_history, event_nums = simulate_events(1, event_size, participants, current_event_number, 
                                           THRESHOLDS["strict"], inactivity_threshold, "strict")
        
        # Accumulate history
        for participant in participants:
            all_participants_history[participant.name].extend(history[participant.name])
            all_rank_history[participant.name].extend(rank_history[participant.name])
        all_event_numbers.extend(event_nums)
        
        if switch_between_reset_modes:
            if (current_event_number) == breakpoint:
                apply_reset(participants,THRESHOLDS["strict"])
                counter += 1
                try:
                    breakpoint += event_distribution[counter]
                    print("Season", counter+1)
                    print("#" * 90)
                except:
                    pass
        current_event_number += 1

    print(f"\nCreating graphs for strict thresholds (Inactivity: {inactivity_threshold})...")
    workbook = create_excel_graphs(all_participants_history, f"strict_{inactivity_threshold}", all_event_numbers, workbook)
    workbook = create_rank_progression_graphs(all_participants_history, all_rank_history, f"strict_{inactivity_threshold}", all_event_numbers, workbook)

    # Save the workbook after all simulations for this threshold
    try:
        # Remove the first empty sheet if it exists
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        
        filename = f'inactivity_threshold_{inactivity_threshold}.xlsx'
        workbook.save(filename)
        print(f"\nSuccessfully saved Excel file: {filename}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        print(f"Current working directory: {os.getcwd()}")

    # After each complete simulation (at the end of the event size loop)
    simulation_results.append(participants.copy())

# Save the single workbook after all simulations are complete
try:
    # Remove the first empty sheet if it exists
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    
    filename = 'volunteer_ranking_simulation_results.xlsx'
    workbook.save(filename)
    print(f"\nSuccessfully saved Excel file: {filename}")
except Exception as e:
    print(f"Error saving Excel file: {e}")
    print(f"Current working directory: {os.getcwd()}")

def create_rank_distribution_excel(all_results, output_file="volunteer_rank_distribution.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rank Distribution"
    
    # Convert results to DataFrame for easier manipulation
    import pandas as pd
    df = pd.DataFrame(all_results)
    df_pivot = df.pivot(index='event_number', columns='name', values='rank_value')
    
    # Write to Excel
    ws['A1'] = "Event"
    for col, name in enumerate(df_pivot.columns, start=2):
        ws.cell(row=1, column=col, value=name)
    
    for row, (idx, data) in enumerate(df_pivot.iterrows(), start=2):
        ws.cell(row=row, column=1, value=idx)
        for col, value in enumerate(data, start=2):
            ws.cell(row=row, column=col, value=value)
    
    # Create chart
    chart = LineChart()
    chart.title = "Volunteer Ranks Over Time"
    chart.style = 13
    chart.x_axis.title = "Event Number"
    chart.y_axis.title = "Rank"
    chart.height = 15
    chart.width = 20
    
    # Set y-axis properties for ranks
    chart.y_axis.scaling.min = 0.5
    chart.y_axis.scaling.max = 4.5
    chart.y_axis.majorUnit = 1
    
    # Simplified axis formatting
    chart.y_axis.majorTickMark = "cross"
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.numFmt = "General"
    chart.y_axis.majorGridlines = None
    
    # Add data series
    data = Reference(ws, min_col=2, min_row=1, max_col=len(df_pivot.columns)+1, 
                    max_row=len(df_pivot)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_pivot)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Style lines
    colors = ['FF0000', '00FF00', '0000FF', 'FFA500', '800080', 
              '008080', 'FF69B4', '4B0082', '006400', 'DC143C']
    
    for i, series in enumerate(chart.series):
        series.smooth = True
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.graphicalProperties.line.width = 25000
        series.graphicalProperties.line.solidFill = colors[i % len(colors)]
    
    # Position legend
    chart.legend.position = 'r'
    chart.plot_area.layout = Layout(
        manualLayout=ManualLayout(
            x=0.1,    # Left position
            y=0.1,    # Top position
            h=0.8,    # Height
            w=0.7     # Width - reduced to make room for legend
        )
    )
    
    # Add chart to worksheet
    ws.add_chart(chart, "A" + str(len(df_pivot) + 5))
    
    wb.save(output_file)

# At the start of your simulation loop, before any events
all_simulation_results = []

# Inside your main event loop, after each event's simulation and rank updates
# (specifically after the simulate_events() call)
event_results = []
for participant in participants:
    event_results.append({
        'name': participant.name,
        'rank_value': {'Platinum': 4, 'Gold': 3, 'Silver': 2, 'Bronze': 1}[participant.rank],
        'event_number': current_event_number
    })
all_simulation_results.extend(event_results)

# Finally, call the function with all results
create_rank_distribution_excel(all_simulation_results)
